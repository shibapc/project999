from aiogram import Bot, Dispatcher, types
from aiogram.utils import executor
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os

from config import BOT_TOKEN

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(bot)
dp.middleware.setup(LoggingMiddleware())

user_data = {}

# Команда /start
@dp.message_handler(commands=['start'])
async def start_command(message: types.Message):
    user_data[message.chat.id] = {
        "sheets": [],
        "current_sheet": None,
        "products": {},
        "quantities": {},
        "step": None
    }
    await message.answer("Привет! Давай начнем создание сметы.\nСколько листов нужно создать (без учета СВОДНОГО)?")

# Получаем количество листов
@dp.message_handler(lambda message: message.chat.id in user_data and user_data[message.chat.id].get("step") is None and message.text.isdigit())
async def get_number_of_sheets(message: types.Message):
    num_sheets = int(message.text)
    user_data[message.chat.id]["sheet_count"] = num_sheets
    user_data[message.chat.id]["step"] = "sheet_names"
    await message.answer("Введи названия листов через запятую:")

# Получаем названия листов и количество каждого МАФа
@dp.message_handler(lambda message: message.chat.id in user_data and user_data[message.chat.id].get("step") == "sheet_names" and "," in message.text)
async def get_sheet_names_and_quantities(message: types.Message):
    sheets = message.text.split(",")
    sheets = [sheet.strip() for sheet in sheets]
    user_data[message.chat.id]["sheets"] = sheets
    user_data[message.chat.id]["step"] = "maf_quantity"
    await ask_maf_quantity(message)

# Спрашиваем количество каждого МАФа
async def ask_maf_quantity(message):
    sheets = user_data[message.chat.id]["sheets"]
    if "quantities" not in user_data[message.chat.id]:
        user_data[message.chat.id]["quantities"] = {}

    if len(user_data[message.chat.id]["quantities"]) < len(sheets):
        next_sheet = sheets[len(user_data[message.chat.id]["quantities"])]
        await message.answer(f"{next_sheet} в каком количестве?")
    else:
        user_data[message.chat.id]["current_sheet"] = sheets[0]
        user_data[message.chat.id]["step"] = "product_name"
        await message.answer(f"Начнем заполнять данные для листа: {sheets[0]}\nВведи название изделия:")

# Получаем количество каждого МАФа
@dp.message_handler(lambda message: message.chat.id in user_data and user_data[message.chat.id].get("step") == "maf_quantity" and message.text.isdigit())
async def get_maf_quantity(message: types.Message):
    sheets = user_data[message.chat.id]["sheets"]
    maf_quantity = int(message.text)
    current_maf = sheets[len(user_data[message.chat.id]["quantities"])]
    user_data[message.chat.id]["quantities"][current_maf] = maf_quantity
    await ask_maf_quantity(message)

# Получаем название изделия
@dp.message_handler(lambda message: message.chat.id in user_data and user_data[message.chat.id].get("step") == "product_name")
async def get_product_name(message: types.Message):
    current_sheet = user_data[message.chat.id]["current_sheet"]
    product_name = message.text
    if current_sheet not in user_data[message.chat.id]["products"]:
        user_data[message.chat.id]["products"][current_sheet] = []

    user_data[message.chat.id]["products"][current_sheet].append({
        "name": product_name,
        "quantity": 0,
        "unit": "",
        "price_per_unit": 0
    })
    user_data[message.chat.id]["step"] = "quantity"
    await message.answer("Введи количество:")

# Получаем количество
@dp.message_handler(lambda message: message.chat.id in user_data and user_data[message.chat.id].get("step") == "quantity" and message.text.isdigit())
async def get_product_quantity(message: types.Message):
    quantity = int(message.text)
    current_sheet = user_data[message.chat.id]["current_sheet"]
    user_data[message.chat.id]["products"][current_sheet][-1]["quantity"] = quantity
    user_data[message.chat.id]["step"] = "unit"
    
    # Создаем клавиатуру с кнопками
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
    buttons = ["м", "м²", "м³", "шт"]
    keyboard.add(*buttons)
    
    await message.answer("Введи единицы измерения (м, м², м³, шт):", reply_markup=keyboard)

# Получаем единицы измерения
@dp.message_handler(lambda message: message.chat.id in user_data and user_data[message.chat.id].get("step") == "unit")
async def get_product_unit(message: types.Message):
    unit = message.text
    current_sheet = user_data[message.chat.id]["current_sheet"]
    
    # Обновляем информацию о единице измерения
    user_data[message.chat.id]["products"][current_sheet][-1]["unit"] = unit
    user_data[message.chat.id]["step"] = "price"  # Переходим к следующему шагу
    
    await message.answer("Введи цену за единицу изделия:", reply_markup=types.ReplyKeyboardRemove())

# Получаем цену за единицу
@dp.message_handler(lambda message: message.chat.id in user_data and user_data[message.chat.id].get("step") == "price" and message.text.replace('.', '', 1).isdigit())
async def get_product_price(message: types.Message):
    price = float(message.text)
    current_sheet = user_data[message.chat.id]["current_sheet"]
    
    # Обновляем цену
    user_data[message.chat.id]["products"][current_sheet][-1]["price_per_unit"] = price
    user_data[message.chat.id]["step"] = "next_product"  # Переход к следующему изделию
    
    await message.answer("Данные сохранены! Введи название следующего изделия или напиши 'далее' для перехода к следующему листу.")

# Обрабатываем ввод следующего изделия или переход к следующему листу
@dp.message_handler(lambda message: message.chat.id in user_data and user_data[message.chat.id].get("step") == "next_product")
async def process_next_product_or_sheet(message: types.Message):
    if message.text.lower() == 'далее':
        current_sheet = user_data[message.chat.id]["current_sheet"]
        sheets = user_data[message.chat.id]["sheets"]
        current_index = sheets.index(current_sheet)
        
        # Если есть следующий лист
        if current_index + 1 < len(sheets):
            next_sheet = sheets[current_index + 1]
            user_data[message.chat.id]["current_sheet"] = next_sheet
            user_data[message.chat.id]["step"] = "product_name"
            await message.answer(f"Переходим к следующему листу: {next_sheet}\nВведи название изделия:")
        else:
            await message.answer("Все данные введены! Сейчас сформирую Excel файл.")
            create_excel(message.chat.id)
            await message.answer("Файл готов! Отправляю...")
            await send_file(message)
    else:
        # Получаем название следующего изделия
        current_sheet = user_data[message.chat.id]["current_sheet"]
        product_name = message.text
        user_data[message.chat.id]["products"][current_sheet].append({
            "name": product_name,
            "quantity": 0,
            "unit": "",
            "price_per_unit": 0
        })
        user_data[message.chat.id]["step"] = "quantity"
        await message.answer("Введи количество:")

# Создание Excel файла
def create_excel(chat_id):
    wb = Workbook()
    sheets = user_data[chat_id]["sheets"]
    alignment = Alignment(horizontal='center', vertical='center')

    for sheet_name in sheets:
        ws = wb.create_sheet(title=sheet_name)
        header = ["Название", "Количество", "Ед. изм.", "Цена за ед.", "Сумма"]
        ws.append(header)
        
        for cell in ws[1]:
            cell.alignment = alignment
        
        total_sum = 0
        for idx, product in enumerate(user_data[chat_id]["products"][sheet_name], start=2):
            formula = f"=B{idx}*D{idx}"
            row = [
                product["name"],
                product["quantity"],
                product["unit"],
                product["price_per_unit"],
                formula
            ]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.alignment = alignment

        # Добавляем строку "ИТОГО"
        total_row_formula = f"=SUM(E2:E{len(user_data[chat_id]['products'][sheet_name]) + 1})"
        total_row = ["ИТОГО", "", "", "", total_row_formula]
        ws.append(total_row)
        for cell in ws[ws.max_row]:
            cell.alignment = alignment

        # Увеличиваем высоту всех строк и задаем выравнивание
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment
            ws.row_dimensions[row[0].row].height = 25

        # Автоматическая подгонка ширины колонок
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Получаем букву колонки
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # Создаем сводный лист
    ws_summary = wb.create_sheet(title="СВОДНЫЙ")
    header = ["№", "МАФ", "Количество", "Итоговая стоимость за 1шт", "Итоговая стоимость за все"]
    ws_summary.append(header)
    
    for cell in ws_summary[1]:
        cell.alignment = alignment

    summary_index = 1
    quantities = user_data[chat_id]["quantities"]
    total_summary_sum_formula_parts = []
    for idx, sheet_name in enumerate(sheets, start=2):
        total_sum_formula = f"=SUM('{sheet_name}'!E2:E{len(user_data[chat_id]['products'][sheet_name]) + 1})"
        total_sum = f"=C{idx}*D{idx}"
        row = [
            summary_index,
            sheet_name,
            quantities[sheet_name],
            total_sum_formula,
            total_sum
        ]
        ws_summary.append(row)
        for cell in ws_summary[ws_summary.max_row]:
            cell.alignment = alignment
        summary_index += 1
        total_summary_sum_formula_parts.append(f"E{idx}")

    # Добавляем строку "ИТОГО" в сводный лист
    total_summary_formula = f"=SUM({','.join(total_summary_sum_formula_parts)})"
    total_summary_row = ["ИТОГО", "", "", "", total_summary_formula]
    ws_summary.append(total_summary_row)
    for cell in ws_summary[ws_summary.max_row]:
        cell.alignment = alignment

    # Увеличиваем высоту всех строк в сводном листе и задаем выравнивание
    for row in ws_summary.iter_rows():
        for cell in row:
            cell.alignment = alignment
        ws_summary.row_dimensions[row[0].row].height = 25

    # Автоматическая подгонка ширины колонок в сводном листе
    for col in ws_summary.columns:
        max_length = 0
        column = col[0].column_letter  # Получаем букву колонки
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_summary.column_dimensions[column].width = adjusted_width

    del wb["Sheet"]  # Удаляем стандартный пустой лист
    file_name = f"smeta_{chat_id}.xlsx"
    wb.save(file_name)

@dp.message_handler(commands=['get_file'])
async def send_file(message: types.Message):
    file_name = f"smeta_{message.chat.id}.xlsx"
    if os.path.exists(file_name):
        await message.answer_document(open(file_name, 'rb'))
    else:
        await message.answer("Файл не найден. Попробуйте создать смету заново.")

if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)