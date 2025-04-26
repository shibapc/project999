import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.merge import MergedCell
from utils.materials_manager import materials_manager
from calculations import calculate_product_cost

logger = logging.getLogger(__name__)

def create_excel(chat_id: int, user_data: dict) -> str:
    """Создаёт Excel-файл со сметой на основе user_data."""
    try:
        # Инициализация рабочей книги
        wb = Workbook()
        wb.remove(wb.active)  # Удаляем стандартный лист

        # Получение данных из user_data
        sheets = user_data[chat_id].get("sheets", [])
        products = user_data[chat_id].get("products", {})
        quantities = user_data[chat_id].get("quantities", {})

        if not sheets:
            logger.error(f"Нет листов для обработки, chat_id={chat_id}")
            return ""

        # Стили для форматирования
        header_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        title_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        title_fill = PatternFill(start_color="A3BFFA", end_color="A3BFFA", fill_type="solid")  # Светло-синий фон

        # Создание папки для смет, если она не существует
        output_dir = "smeta_files"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # Создание листов
        for sheet_name in sheets:
            ws = wb.create_sheet(title=sheet_name)
            maf_quantity = quantities.get(sheet_name, 1)

            # # Добавление логотипа перед таблицей
            # from openpyxl.drawing.image import Image
            # 
            # # Путь к файлу логотипа (замените на ваш путь)
            # logo_path = "path/to/logo.png"
            # 
            # # Проверка существования файла
            # if os.path.exists(logo_path):
            #     img = Image(logo_path)
            #     # Установка размера изображения (опционально, в пикселях)
            #     img.width = 150  # Ширина изображения
            #     img.height = 50  # Высота изображения
            #     # Добавление изображения в ячейку A1
            #     ws.add_image(img, "A1")
            # else:
            #     logger.warning(f"Файл логотипа не найден: {logo_path}")

            # Заголовок листа
            ws.append([f"{sheet_name}"])
            title_cell = ws.cell(row=1, column=1)
            title_cell.font = title_font
            title_cell.alignment = center_align
            title_cell.fill = title_fill
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
            ws.row_dimensions[1].height = 30

            # Заголовки таблицы
            row = 2
            headers = [
                "№",
                "Наименование",
                "Параметры",
                "",
                "",
                "",
                "",
                "калькуляция",
                "",
                "",
                ""
            ]
            ws.append(headers)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
                cell.fill = header_fill
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
            ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=11)
            row += 1

            sub_headers = [
                "", "",
                "Длина, мм",
                "Ширина, мм",
                "Высота, мм",
                "Кол-во, шт.",
                "",
                "статья затрат",
                "ед.изм.",
                "расход",
                "сумма"
            ]
            ws.append(sub_headers)
            for col, header in enumerate(sub_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
                cell.fill = header_fill
            row += 1

            # Получение продуктов для листа
            sheet_products = products.get(sheet_name, [])
            if not sheet_products:
                logger.warning(f"Нет продуктов для листа '{sheet_name}', chat_id={chat_id}")
                continue

            # Инициализация счетчиков
            item_index = 1
            total_cost = 0.0
            total_transport_cost = 0.0
            total_installation_cost = 0.0

            # Обработка изделий
            for product in sheet_products:
                category = product.get("category")
                if not category:
                    logger.error(f"Отсутствует категория для продукта: {product}, chat_id={chat_id}")
                    continue

                if category in ["Доставка", "Работы"]:
                    continue  # Транспорт и монтаж обрабатываются отдельно

                section = materials_manager.get_category_key(category)
                item = materials_manager.get_item(product.get("name", ""), section)
                if not item:
                    logger.error(f"Элемент '{product.get('name', '')}' не найден для категории '{category}', chat_id={chat_id}")
                    continue

                quantity = product.get("quantity", 1) * maf_quantity
                params = {
                    **{p["key"]: product.get(p["key"], 0) for p in item.get("parameters", [])},
                    "quantity": quantity,
                    "all_products": sheet_products
                }
                cost_data = calculate_product_cost(item, params, quantity)

                start_row = row
                product_cost = 0.0

                if category == "Изделия" and item.get("calculation", {}).get("type") == "complex":
                    # Подсчет количества строк для материалов и работ
                    materials = item.get("calculation", {}).get("materials", [])
                    works = item.get("calculation", {}).get("works", [])
                    total_rows = len(materials) + len(works)

                    # Добавление самого изделия
                    width = product.get("width_mm", "")
                    length = product.get("length_mm", "")
                    height = product.get("height_mm", "")
                    ws.append([
                        item_index,
                        product.get("name", ""),
                        length,
                        width,
                        height,
                        quantity,
                        "",
                        "", "", "", ""
                    ])
                    for col in range(1, 12):
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        cell.alignment = center_align if col != 2 else left_align
                    row += 1

                    # Обработка материалов
                    details = cost_data.get("детали", {})
                    for mat in materials:
                        mat_name = mat["name"]
                        mat_quantity = details.get(f"материал_{mat_name}", {}).get("количество", 0)
                        mat_cost = details.get(f"материал_{mat_name}", {}).get("стоимость", 0)
                        mat_item = materials_manager.get_item(mat_name, "materials")
                        if not mat_item:
                            logger.error(f"Материал '{mat_name}' не найден, chat_id={chat_id}")
                            continue

                        ws.append([
                            "",
                            f" - {mat_name}",
                            "", "", "", "",
                            mat_name,
                            mat_item.get("unit", "ед"),
                            mat_quantity,
                            mat_cost / mat_quantity if mat_quantity else 0,
                            mat_cost
                        ])
                        for col in range(1, 12):
                            cell = ws.cell(row=row, column=col)
                            cell.border = border
                            cell.alignment = center_align if col != 2 else left_align
                        product_cost += mat_cost
                        row += 1

                    # Обработка работ
                    for work in works:
                        work_name = work["name"]
                        work_quantity = details.get(f"работа_{work_name}", {}).get("количество", 0)
                        work_cost = details.get(f"работа_{work_name}", {}).get("стоимость", 0)
                        work_item = materials_manager.get_item(work_name, "works")
                        if not work_item:
                            logger.error(f"Работа '{work_name}' не найдена, chat_id={chat_id}")
                            continue

                        ws.append([
                            "",
                            f" - {work_name}",
                            "", "", "", "",
                            work_name,
                            work_item.get("unit", "ед"),
                            work_quantity,
                            work_cost / work_quantity if work_quantity else 0,
                            work_cost
                        ])
                        for col in range(1, 12):
                            cell = ws.cell(row=row, column=col)
                            cell.border = border
                            cell.alignment = center_align if col != 2 else left_align
                        product_cost += work_cost
                        row += 1

                    # Объединение ячеек для названия изделия
                    if total_rows > 0:
                        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row + total_rows - 1, end_column=2)
                        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + total_rows - 1, end_column=1)

                else:
                    # Обычные материалы
                    width = product.get("width_mm", "")
                    length = product.get("length_mm", "")
                    height = product.get("height_mm", "")
                    unit = product.get("unit", "ед")
                    price_per_unit = product.get("price_per_unit", 0)
                    total_product_cost = product.get("total_cost", 0) * maf_quantity

                    ws.append([
                        item_index,
                        product.get("name", ""),
                        length,
                        width,
                        height,
                        quantity,
                        product.get("name", ""),
                        unit,
                        quantity,
                        price_per_unit,
                        total_product_cost
                    ])
                    for col in range(1, 12):
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        cell.alignment = center_align if col != 2 else left_align
                    product_cost = total_product_cost
                    row += 1

                # Итог за штуку
                ws.append(["", "Итого за штуку", "", "", "", "", "", "", "", "", product_cost])
                for col in range(1, 12):
                    cell = ws.cell(row=row, column=col)
                    cell.font = bold_font
                    cell.border = border
                    cell.alignment = center_align if col != 2 else left_align
                    if col in [2, 11]:
                        cell.fill = title_fill
                total_cost += product_cost
                row += 1
                item_index += 1

            # Таблица транспортных расходов
            ws.append(["Транспортные расходы"])
            category_cell = ws.cell(row=row, column=1)
            category_cell.font = bold_font
            category_cell.fill = title_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
            row += 1

            ws.append(sub_headers)
            for col, header in enumerate(sub_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
                cell.fill = header_fill
            row += 1

            transport_index = 1
            for product in sheet_products:
                if product.get("category") != "Доставка":
                    continue
                ws.append([
                    transport_index,
                    product.get("name", ""),
                    "", "", "", product.get("quantity", 1) * maf_quantity,
                    product.get("name", ""),
                    product.get("unit", "ед"),
                    product.get("quantity", 1) * maf_quantity,
                    product.get("price_per_unit", 0),
                    product.get("total_cost", 0) * maf_quantity
                ])
                for col in range(1, 12):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = center_align if col != 2 else left_align
                total_transport_cost += product.get("total_cost", 0) * maf_quantity
                transport_index += 1
                row += 1

            ws.append(["", "Итого транспортные расходы", "", "", "", "", "", "", "", "", total_transport_cost])
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.font = bold_font
                cell.border = border
                cell.alignment = center_align if col != 2 else left_align
                if col in [2, 11]:
                    cell.fill = title_fill
            row += 1

            # Таблица монтажных работ
            ws.append(["Монтажные работы"])
            category_cell = ws.cell(row=row, column=1)
            category_cell.font = bold_font
            category_cell.fill = title_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
            row += 1

            ws.append(sub_headers)
            for col, header in enumerate(sub_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
                cell.fill = header_fill
            row += 1

            installation_index = 1
            for product in sheet_products:
                if product.get("category") != "Работы":
                    continue
                ws.append([
                    installation_index,
                    product.get("name", ""),
                    "", "", "", product.get("quantity", 1) * maf_quantity,
                    product.get("name", ""),
                    product.get("unit", "ед"),
                    product.get("quantity", 1) * maf_quantity,
                    product.get("price_per_unit", 0),
                    product.get("total_cost", 0) * maf_quantity
                ])
                for col in range(1, 12):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = center_align if col != 2 else left_align
                total_installation_cost += product.get("total_cost", 0) * maf_quantity
                installation_index += 1
                row += 1

            ws.append(["", "Итого монтажных работ", "", "", "", "", "", "", "", "", total_installation_cost])
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.font = bold_font
                cell.border = border
                cell.alignment = center_align if col != 2 else left_align
                if col in [2, 11]:
                    cell.fill = title_fill
            row += 1

            # Итоговая себестоимость
            final_cost = total_cost + total_transport_cost + total_installation_cost
            ws.append(["", "ИТОГО ПОЛНАЯ СЕБЕСТОИМОСТЬ", "", "", "", "", "", "", "", "", final_cost])
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.font = bold_font
                cell.border = border
                cell.alignment = center_align if col != 2 else left_align
                if col in [2, 11]:
                    cell.fill = title_fill

            # Автоподбор ширины столбцов
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column = get_column_letter(col_idx)
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value and not isinstance(cell, MergedCell):
                            try:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                            except:
                                pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

        # Сохранение файла в smeta_files/
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(output_dir, f"estimate_{chat_id}_{timestamp}.xlsx")
        wb.save(filename)
        logger.info(f"Excel-файл создан: {filename}, chat_id={chat_id}")
        return filename

    except Exception as e:
        logger.error(f"Ошибка при создании Excel-файла: {e}, chat_id={chat_id}")
        return ""