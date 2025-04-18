import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Alignment

# Настройка логирования
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)


def create_excel(chat_id: int, user_data: dict):
    """Создаёт Excel-файл с данными сметы."""
    try:
        logger.debug(f"Создание Excel для chat_id={chat_id}")
        wb = Workbook()
        sheets = user_data[chat_id]["sheets"]
        if not sheets:
            logger.error("Список sheets пустой")
            raise ValueError("Список sheets пустой")

        alignment = Alignment(horizontal="center", vertical="center")

        for sheet_name in sheets:
            ws = wb.create_sheet(title=sheet_name)
            row_offset = 1

            # Группировка продуктов по категориям
            products = user_data[chat_id]["products"].get(sheet_name, [])
            categories = set(product["category"] for product in products)

            for category in categories:
                ws.append([f"{category}"])
                row_offset += 1
                header = [
                    "Код",
                    "Название",
                    "Ширина (мм)",
                    "Высота (мм)",
                    "Объём (м³)",
                    "Количество",
                    "Ед. изм.",
                    "Цена за ед.",
                    "Сумма",
                ]
                ws.append(header)
                row_offset += 1

                for cell in ws[row_offset - 1]:
                    cell.alignment = alignment

                category_products = [p for p in products if p["category"] == category]
                category_start_row = row_offset

                for idx, product in enumerate(category_products, start=1):
                    formula = f"=F{row_offset}*H{row_offset}"
                    width = product.get("width", "")
                    height = product.get("height", "")
                    volume = product.get("volume_m3", "")
                    row = [
                        product["id"],
                        product["name"],
                        width,
                        height,
                        volume,
                        product["quantity"],
                        product["unit"],
                        product["price_per_unit"],
                        formula,
                    ]
                    ws.append(row)
                    row_offset += 1
                    for cell in ws[ws.max_row]:
                        cell.alignment = alignment

                total_row_formula = f"=SUM(I{category_start_row}:I{row_offset-1})"
                ws.append(
                    ["ИТОГО " + category, "", "", "", "", "", "", "", total_row_formula]
                )
                row_offset += 1
                for cell in ws[ws.max_row]:
                    cell.alignment = alignment

                ws.append([])  # Пустая строка между категориями
                row_offset += 1

            # Форматирование
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = alignment
                ws.row_dimensions[row[0].row].height = 25

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width

        # Сводный лист
        ws_summary = wb.create_sheet(title="СВОДНЫЙ")
        header = ["№", "МАФ", "Количество", "Материалы", "Работы", "Прочее", "Итого"]
        ws_summary.append(header)

        for cell in ws_summary[1]:
            cell.alignment = alignment

        summary_index = 1
        quantities = user_data[chat_id]["quantities"]
        total_summary = {"Материалы": 0, "Работы": 0, "Прочее": 0}

        for idx, sheet_name in enumerate(sheets, start=2):
            products = user_data[chat_id]["products"].get(sheet_name, [])
            category_sums = {"Материалы": 0, "Работы": 0, "Прочее": 0}

            for product in products:
                sum_value = product["quantity"] * product["price_per_unit"]
                category_sums[product["category"]] += sum_value

            total_sum = sum(category_sums.values())
            row = [
                summary_index,
                sheet_name,
                quantities.get(sheet_name, 0),
                category_sums["Материалы"],
                category_sums["Работы"],
                category_sums["Прочее"],
                total_sum,
            ]
            ws_summary.append(row)
            row_offset += 1
            for cell in ws_summary[ws_summary.max_row]:
                cell.alignment = alignment
            summary_index += 1

            for cat in total_summary:
                total_summary[cat] += category_sums[cat]

        total_summary_row = [
            "ИТОГО",
            "",
            "",
            total_summary["Материалы"],
            total_summary["Работы"],
            total_summary["Прочее"],
            sum(total_summary.values()),
        ]
        ws_summary.append(total_summary_row)
        for cell in ws_summary[ws.max_row]:
            cell.alignment = alignment

        for row in ws_summary.iter_rows():
            for cell in row:
                cell.alignment = alignment
            ws_summary.row_dimensions[row[0].row].height = 25

        for col in ws_summary.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws_summary.column_dimensions[column].width = adjusted_width

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Исправление пути сохранения
        smeta_dir = os.path.join(os.getcwd(), "smeta_files")
        os.makedirs(smeta_dir, exist_ok=True)
        file_name = os.path.join(smeta_dir, f"smeta_{chat_id}.xlsx")
        logger.debug(f"Сохранение файла: {os.path.abspath(file_name)}")
        wb.save(file_name)
        if not os.path.exists(file_name):
            logger.error(f"Файл {file_name} не был создан")
            raise FileNotFoundError(f"Файл {file_name} не был создан")
        logger.info(f"Файл {file_name} успешно создан")
    except Exception as e:
        logger.error(f"Ошибка при создании Excel: {str(e)}")
        raise
