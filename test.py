import unittest
import json
import os
import asyncio
from unittest.mock import AsyncMock, patch, MagicMock
from openpyxl import load_workbook
from utils.materials_manager import MaterialsManager
from calculations import calculate_product_cost
from handlers.manual import (
    validate_and_save_param,
    handle_param_input,
    format_calculation_result,
    calculate_product_cost_wrapper,
    get_product_quantity,
)
from utils.excel_creator import create_excel

class TestConstructionCalculator(unittest.TestCase):
    def setUp(self):
        """Инициализация перед каждым тестом."""
        # Создание временного materials.json
        self.test_json = {
            "categories": [
                {"name": "Материалы", "key": "materials", "phase": "material"},
                {"name": "Изделия", "key": "templates", "phase": "material"},
                {"name": "Работы", "key": "works", "phase": "non_material"},
                {"name": "Доставка", "key": "other", "phase": "non_material"},
            ],
            "materials": [
                {
                    "id": 1,
                    "name": "Доска",
                    "category": "Материалы",
                    "unit": "шт",
                    "base_price": 4000,
                },
                {
                    "id": 14,
                    "name": "Металлическая труба",
                    "category": "Материалы",
                    "unit": "шт",
                    "variable": True,
                    "parameters": [
                        {
                            "name": "длина",
                            "key": "length_mm",
                            "type": "float",
                            "min": 1,
                            "max": 10000,
                            "prompt": "Укажи длину (мм):",
                        },
                        {
                            "name": "диаметр",
                            "key": "diameter_mm",
                            "type": "float",
                            "min": 10,
                            "max": 1000,
                            "prompt": "Укажи диаметр (мм):",
                        },
                    ],
                    "calculation": {
                        "type": "volume",
                        "volume_formula": "math.pi * (diameter_mm / 2) ** 2 * length_mm / 1000000000",
                        "cost_per_m3": 1000000,
                    },
                },
            ],
            "templates": [
                {
                    "id": 2,
                    "name": "Скамейка",
                    "category": "Изделия",
                    "unit": "шт",
                    "calculation": {
                        "type": "complex",
                        "materials": [
                            {"name": "Доска", "quantity_per_unit": 5},
                        ],
                        "works": [
                            {"name": "Сварка", "quantity_per_unit": 2},
                        ],
                    },
                },
            ],
            "works": [
                {
                    "id": 3,
                    "name": "Сварка",
                    "category": "Работы",
                    "unit": "шт",
                    "base_price": 1000,
                },
            ],
            "other": [
                {
                    "id": 4,
                    "name": "Доставка",
                    "category": "Доставка",
                    "unit": "шт",
                    "calculation": {
                        "type": "price_formula",
                        "price_formula": "5000 if sum_material_volume < 1 else 10000",
                    },
                },
            ],
        }
        with open("test_materials.json", "w", encoding="utf-8") as f:
            json.dump(self.test_json, f, ensure_ascii=False, indent=2)

        # Инициализация MaterialsManager
        self.materials_manager = MaterialsManager("test_materials.json")

        # Моки для Telegram
        self.update = AsyncMock()
        self.update.message.chat_id = 12345
        self.context = AsyncMock()
        self.user_data = {
            12345: {
                "products": {"Скамейка": []},
                "current_sheet": "Скамейка",
                "sheets": ["Скамейка"],
                "quantities": {"Скамейка": 2},
                "material_phase": "material",
                "awaiting_param": None,
                "awaiting_quantity": False,
                "awaiting_price": False,
                "has_non_material": False,
            }
        }

    def tearDown(self):
        """Очистка после каждого теста."""
        if os.path.exists("test_materials.json"):
            os.remove("test_materials.json")
        for file in os.listdir():
            if file.startswith("estimate_12345_"):
                os.remove(file)

    def test_materials_manager_load(self):
        """Тест загрузки и валидации materials.json."""
        self.assertEqual(len(self.materials_manager.db["materials"]), 2)
        self.assertEqual(len(self.materials_manager.db["categories"]), 4)
        item = self.materials_manager.get_item("Металлическая труба", "materials")
        self.assertIsNotNone(item)
        self.assertEqual(item["name"], "Металлическая труба")

    def test_calculate_product_cost_base_price(self):
        """Тест расчёта для base_price."""
        item = self.materials_manager.get_item("Доска", "materials")
        params = {"quantity": 5}
        result = calculate_product_cost(item, params, 5)
        self.assertEqual(result["общая_стоимость"], 20000)
        self.assertEqual(result["детали"], {})

    def test_calculate_product_cost_volume(self):
        """Тест расчёта для volume (металлическая труба)."""
        item = self.materials_manager.get_item("Металлическая труба", "materials")
        params = {"length_mm": 2000, "diameter_mm": 100, "quantity": 1}
        result = calculate_product_cost(item, params, 1)
        volume = 3.14159 * (100 / 2) ** 2 * 2000 / 1000000000  # π * (d/2)^2 * l
        expected_cost = volume * 1000000
        self.assertAlmostEqual(result["общая_стоимость"], expected_cost, places=2)
        self.assertAlmostEqual(result["детали"]["объём_м3"], volume, places=5)

    def test_calculate_product_cost_complex(self):
        """Тест расчёта для complex (скамейка)."""
        item = self.materials_manager.get_item("Скамейка", "templates")
        params = {"quantity": 1}
        result = calculate_product_cost(item, params, 1)
        expected_cost = (5 * 4000) + (2 * 1000)  # 5 досок + 2 сварки
        self.assertEqual(result["общая_стоимость"], expected_cost)
        self.assertEqual(result["детали"]["материал_Доска"]["количество"], 5)
        self.assertEqual(result["детали"]["работа_Сварка"]["количество"], 2)

    def test_calculate_product_cost_price_formula(self):
        """Тест расчёта для price_formula (доставка)."""
        item = self.materials_manager.get_item("Доставка", "other")
        params = {
            "quantity": 1,
            "all_products": [
                {"объём_м3": 0.1},
                {"объём_м3": 0.2},
            ],
        }
        result = calculate_product_cost(item, params, 1)
        self.assertEqual(result["общая_стоимость"], 5000)  # sum_material_volume = 0.3 < 1

    async def test_validate_and_save_param(self):
        """Тест валидации параметров."""
        item = self.materials_manager.get_item("Металлическая труба", "materials")
        product = {"name": "Металлическая труба", "category": "Материалы"}
        param = item["parameters"][0]  # длина
        self.update.message.text = "2000"
        result = await validate_and_save_param(
            self.update, self.user_data, product, item, param, "2000"
        )
        self.assertTrue(result)
        self.assertEqual(product["length_mm"], 2000)

        self.update.message.text = "0"
        result = await validate_and_save_param(
            self.update, self.user_data, product, item, param, "0"
        )
        self.assertFalse(result)

    async def test_handle_param_input(self):
        """Тест обработки ввода параметров."""
        item = self.materials_manager.get_item("Металлическая труба", "materials")
        product = {"name": "Металлическая труба", "category": "Материалы"}
        self.user_data[12345]["awaiting_param"] = item["parameters"][0]  # длина
        self.update.message.text = "2000"
        state = await handle_param_input(
            self.update, self.context, self.user_data, product, item
        )
        self.assertEqual(state, 5)  # Ожидаем следующий параметр
        self.assertEqual(self.user_data[12345]["awaiting_param"]["name"], "диаметр")

    async def test_format_calculation_result(self):
        """Тест форматирования результатов."""
        item = self.materials_manager.get_item("Металлическая труба", "materials")
        product = {
            "name": "Металлическая труба",
            "category": "Материалы",
            "length_mm": 2000,
            "diameter_mm": 100,
        }
        cost_data = {
            "общая_стоимость": 15707.95,
            "детали": {"объём_м3": 0.01570795},
        }
        result = await format_calculation_result(product, item, cost_data)
        expected = (
            "Для 'Металлическая труба' (длина: 2000, диаметр: 100):\n"
            "Объём: 0.016 м³\n"
            "Розничная стоимость: 15708 ₽\n"
        )
        self.assertEqual(result, expected)

    async def test_get_product_quantity(self):
        """Тест обработки количества."""
        item = self.materials_manager.get_item("Доска", "materials")
        product = {
            "name": "Доска",
            "category": "Материалы",
            "unit": "шт",
        }
        self.user_data[12345]["products"]["Скамейка"].append(product)
        self.user_data[12345]["awaiting_quantity"] = True
        self.update.message.text = "5"
        state = await get_product_quantity(self.update, self.context, self.user_data)
        self.assertEqual(state, 6)
        self.assertEqual(product["quantity"], 5)
        self.assertEqual(product["total_cost"], 20000)

    def test_create_excel(self):
        """Тест генерации Excel-файла."""
        self.user_data[12345]["products"]["Скамейка"] = [
            {
                "name": "Доска",
                "category": "Материалы",
                "unit": "шт",
                "quantity": 5,
                "price_per_unit": 4000,
                "total_cost": 20000,
                "объём_м3": 0.1,
            },
            {
                "name": "Сварка",
                "category": "Работы",
                "unit": "шт",
                "quantity": 2,
                "price_per_unit": 1000,
                "total_cost": 2000,
            },
        ]
        filename = create_excel(12345, self.user_data)
        self.assertTrue(os.path.exists(filename))

        wb = load_workbook(filename)
        ws = wb["Скамейка"]
        self.assertEqual(ws["A1"].value, "Категория")
        self.assertEqual(ws["A2"].value, "Материалы")
        self.assertEqual(ws["B2"].value, "Доска")
        self.assertEqual(ws["C2"].value, 10)  # quantity * maf_quantity (5 * 2)
        self.assertEqual(ws["F2"].value, 40000)  # total_cost * maf_quantity
        self.assertEqual(ws["G2"].value, 0.2)  # объём_м3 * maf_quantity
        self.assertEqual(ws["E4"].value, "Итого:")
        self.assertEqual(ws["F4"].value, 44000)  # (20000 + 2000) * 2

    def test_dynamic_categories(self):
        """Тест работы с динамическими категориями."""
        categories = self.materials_manager.get_categories_by_phase("material")
        self.assertEqual(len(categories), 2)
        self.assertEqual(categories[0]["name"], "Материалы")
        self.assertEqual(self.materials_manager.get_category_key("Материалы"), "materials")

def async_test(f):
    """Декоратор для асинхронных тестов."""
    def wrapper(*args, **kwargs):
        loop = asyncio.get_event_loop()
        return loop.run_until_complete(f(*args, **kwargs))
    return wrapper

# Применение декоратора к асинхронным тестам
TestConstructionCalculator.test_validate_and_save_param = async_test(
    TestConstructionCalculator.test_validate_and_save_param
)
TestConstructionCalculator.test_handle_param_input = async_test(
    TestConstructionCalculator.test_handle_param_input
)
TestConstructionCalculator.test_format_calculation_result = async_test(
    TestConstructionCalculator.test_format_calculation_result
)
TestConstructionCalculator.test_get_product_quantity = async_test(
    TestConstructionCalculator.test_get_product_quantity
)

if __name__ == "__main__":
    unittest.main()